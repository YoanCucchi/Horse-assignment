import random
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from collections import defaultdict
import openpyxl

global semaines
# un cavalier n'aura pas le meme cheval pendant X semaines
semaines = 6

def ajouter_manuellement():
	historique = lire_historique(historique_file)
	cavalier = simpledialog.askstring("Ajouter manuellement", "Entrez le nom du cavalier:")
	cheval = simpledialog.askstring("Ajouter manuellement", "Entrez le nom du cheval:")
	
	if cavalier and cheval:
		if cavalier not in historique:
			historique[cavalier] = [cheval]
		else:
			historique[cavalier].append(cheval)
		messagebox.showinfo("Succès", f"{cheval} a été ajouté manuellement pour {cavalier} dans l'historique.")
		with open(historique_file, 'w') as f:
			for cavalier, chevaux in historique.items():
				f.write(f'{cavalier}:{",".join(chevaux)}\n')
	else:
		messagebox.showerror("Erreur", "Veuillez entrer le nom du cavalier et du cheval.")

def lire_fichier(file):
	with open(file, 'r') as f:
		content = f.readlines()
	content = [x.strip().split(',') for x in content if x.strip()]
	return content

def taille_compatibles(taille_cheval, taille_cavalier):
	if taille_cavalier == 'Petit':
		if taille_cheval == 'Petit':
			return True
	elif taille_cavalier == 'Moyen':
		if taille_cheval == "Moyen":
			return True
	elif taille_cavalier == 'Moyen/Grand':
		if taille_cheval == "Moyen/Grand" or taille_cheval == "Moyen":
			return True
	elif taille_cavalier == 'Grand':
		if taille_cheval == 'Grand' or taille_cheval == 'Moyen/Grand':
			return True
	return False

def candidats_pour_cavalier(cavalier, semaines_restriction):
	_, cavalier_taille, _, _ = cavalier
	candidats = []

	for cheval in chevaux:
		cheval_nom, cheval_taille = cheval
		if taille_compatibles(cheval_taille, cavalier_taille):
			candidats.append(cheval_nom)

	for cheval_precedent in historique.get(cavalier[0], [])[-semaines_restriction:]:
		if cheval_precedent in candidats:
			candidats.remove(cheval_precedent)
	return candidats

def attribution(chevaux, cavaliers, historique):
	attribution_courante = {}
	cavaliers_restants = cavaliers.copy()

	while cavaliers_restants:
		candidats_par_cavalier = [(cavalier, candidats_pour_cavalier(cavalier, semaines)) for cavalier in cavaliers_restants]
		cavaliers_tries = sorted(candidats_par_cavalier, key=lambda x: len(x[1]))
		cavalier, candidats = cavaliers_tries[0]
		cavalier_nom, cavalier_taille, cavalier_jour, cavalier_créneau = cavalier

		if candidats:
			cheval_attribue = random.choice(candidats)
			attribution_courante[cavalier_nom, cavalier_jour] = cheval_attribue, cavalier_créneau
			chevaux.remove([cheval for cheval in chevaux if cheval[0] == cheval_attribue][0])
			cavaliers_restants.remove(cavalier)
		else:
			candidats_semaines_moins_1 = candidats_pour_cavalier(cavalier, semaines - 1)
			candidats_semaines_moins_2 = candidats_pour_cavalier(cavalier, semaines - 2)
			if candidats_semaines_moins_1:
				cheval_attribue = random.choice(candidats_semaines_moins_1)
				attribution_courante[cavalier_nom, cavalier_jour] = cheval_attribue, cavalier_créneau
				chevaux.remove([cheval for cheval in chevaux if cheval[0] == cheval_attribue][0])
				cavaliers_restants.remove(cavalier)
			elif candidats_semaines_moins_2:
				cheval_attribue = random.choice(candidats_semaines_moins_2)
				attribution_courante[cavalier_nom, cavalier_jour] = cheval_attribue, cavalier_créneau
				chevaux.remove([cheval for cheval in chevaux if cheval[0] == cheval_attribue][0])
				cavaliers_restants.remove(cavalier)
			else:
				messagebox.showerror("Erreur", "AUCUN CHEVAL ATTRIBUE pour {} !!!!".format(cavalier_nom))
				cavaliers_restants.remove(cavalier)

	return attribution_courante

def sauvegarder_historique(historique, attribution_courante):
	for cavalier, cheval in attribution_courante.items():
		if cavalier[0] not in historique:
			historique[cavalier[0]] = [cheval[0]]
		else:
			historique[cavalier[0]].append(cheval[0])

	with open('historique.txt', 'w') as f:
		for cavalier, chevaux in historique.items():
			f.write(f'{cavalier}:{",".join(chevaux)}\n')

chevaux = lire_fichier('chevaux.txt')
cavaliers = lire_fichier('cavaliers.txt')

def lire_historique(file):
	with open(file, 'r') as f:
		content = f.readlines()
	content = [x.strip().split(':') for x in content if x.strip()]
	return {line[0]: line[1].split(',') for line in content if len(line) >= 2}

def browse_file(type):
	file_path = filedialog.askopenfilename()
	return file_path

def load_cavaliers():
	global cavaliers
	cavaliers = lire_fichier(cavaliers_file)
	load_button.destroy()
	display_cavaliers()

def display_cavaliers():
	global displayed_cavaliers
	for widget in cavalier_frame.winfo_children():
		widget.destroy()

	cavaliers_by_jour = defaultdict(list)
	for cavalier in cavaliers:
		nom, taille, jour, créneau = cavalier
		cavaliers_by_jour[jour].append(cavalier)

	jours_order = ['Mercredi', 'Samedi']
	sorted_jours = sorted(cavaliers_by_jour.keys(), key=lambda x: jours_order.index(x))

	row_index = 0
	displayed_cavaliers = []
	for jour_col, jour in enumerate(sorted_jours):
		cavalier_list = cavaliers_by_jour[jour]

		cavaliers_by_créneau = defaultdict(list)
		for cavalier in cavalier_list:
			nom, taille, jour, créneau = cavalier
			cavaliers_by_créneau[créneau].append(cavalier)

		sorted_créneaux = sorted(cavaliers_by_créneau.keys())

		max_rows = 0
		for créneau in sorted_créneaux:
			cavalier_list = cavaliers_by_créneau[créneau]
			tk.Label(cavalier_frame, text=f"Jour: {jour}, Créneau: {créneau}", font=("Arial", 14, "bold")).grid(row=row_index, column=jour_col * 9, columnspan=8, sticky='nsew')
			row_index += 1

			cavalier_list = sorted(cavalier_list, key=lambda x: x[0])
			i = 0
			for cavalier in cavalier_list:
				nom, taille, jour, créneau = cavalier
				check_var = tk.BooleanVar()
				if jour == "Mercredi":
					sticky_value = 'w'
				else:
					sticky_value = 'e'
				check_button = tk.Checkbutton(cavalier_frame, text=f"{nom}", variable=check_var)
				check_button.grid(row=row_index + i // 8, column=jour_col * 9 + i % 8, sticky=sticky_value, padx=0)
				cavalier_vars.append(check_var)
				displayed_cavaliers.append(cavalier)
				i += 1
			row_index += (i + 7) // 8
			max_rows = max(max_rows, row_index)

		# Ajout d'une ligne verticale entre les colonnes Mercredi et Samedi
		if jour_col < len(sorted_jours) - 1:
			vertical_line = tk.Canvas(cavalier_frame, bg="gray", width=2, height=450)
			vertical_line.grid(row=0, column=(jour_col + 1) * 9 - 1, rowspan=100)
		row_index = 0

def main():
	global chevaux, cavaliers, historique
	chevaux = lire_fichier(chevaux_file)

	try:
		historique = lire_historique(historique_file)
	except FileNotFoundError:
		historique = {}


	check_button = tk.Checkbutton(cavalier_frame, text=f"{nom}", variable=check_var)
	check_button.grid(row=row_index + i // 8, column=jour_col * 9 + i % 8, sticky=sticky_value, padx=0)
 
	cavaliers_a_traiter = [c for c, var in zip(displayed_cavaliers, cavalier_vars) if var.get()]
	print(cavaliers_a_traiter)
	attribution_courante = attribution(chevaux, cavaliers_a_traiter, historique)
	sauvegarder_noms_cavaliers(attribution_courante)
	display_attribution(attribution_courante)
	sauvegarder_historique(historique, attribution_courante)

def display_attribution(attribution_courante):
	tree.delete(*tree.get_children())
	for cavalier, cheval in attribution_courante.items():
		tree.insert("", "end", values=(cavalier[0], cheval[0]))

	for var in cavalier_vars:
		var.set(False)

def sauvegarder_noms_cavaliers(attribution_courante):
	# Charger le fichier Excel existant
	wb = openpyxl.load_workbook('feuille de monte.xlsx')

	# Sélectionner la feuille de calcul à modifier
	for cavalier, cheval in attribution_courante.items():
		if "Mercredi" in cavalier[1]:
			sheet = wb['Mercredi']
		else:
			sheet = wb['Samedi']
		# Trouver la ligne et la colonne correspondantes
		for row in range(1, sheet.max_row+1):
			if sheet.cell(row=row, column=1).value == cheval[0]:
				for col in range(1, sheet.max_column+1):
				#print(f'{sheet.cell(row=row, column=col).value}')
					if sheet.cell(row=1, column=col).value == cheval[1]:
						# Écrire le nom du cavalier dans la cellule correspondante
						sheet.cell(row=row, column=col).value = cavalier[0]
	# Enregistrer les modifications dans le fichier Excel
	wb.save('feuille de monte.xlsx')

app = tk.Tk()
app.title("Attribution Chevaux-Cavaliers")

chevaux_file = 'chevaux.txt'
cavaliers_file = 'cavaliers.txt'
historique_file = 'historique.txt'

cavalier_vars = []
displayed_cavaliers = []

load_button = tk.Button(app, text="Charger", command=load_cavaliers)
load_button.grid(row=4, column=3)
tk.Button(app, text="Ajouter manuellement", command=ajouter_manuellement).grid(row=6, column=4, columnspan=4, sticky="we")

app.columnconfigure(0, weight=1, minsize=0)
app.columnconfigure(1, weight=0, minsize=0)
app.columnconfigure(2, weight=0, minsize=0)
app.columnconfigure(3, weight=1, minsize=0)
app.columnconfigure(4, weight=1, minsize=0)
app.columnconfigure(5, weight=0, minsize=0)
app.columnconfigure(6, weight=0, minsize=0)
app.columnconfigure(7, weight=0, minsize=0)

cavalier_frame = tk.Frame(app)
cavalier_frame.configure(height=500)
cavalier_frame.grid(row=4, column=0, columnspan=5, pady=0, padx=0)

tk.Button(app, text="Attribuer", command=main).grid(row=5, column=2, columnspan=2, sticky="we")

tree = ttk.Treeview(app, columns=("Cavalier", "Cheval"), show="headings", height=15)
tree.heading("Cavalier", text="Cavalier")
tree.heading("Cheval", text="Cheval")
tree.grid(row=6, column=0, columnspan=5, pady=1)

app.protocol("WM_DELETE_WINDOW", app.quit)
app.mainloop()